import datetime
from decimal import Decimal
from openpyxl import load_workbook
from django.core.management.base import BaseCommand, CommandError
from contracts.models import Contract

class Command(BaseCommand):
    help = "Importa los datos de contratos desde un archivo Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='C:\\Users\\edwin.paz\\Documents\\Tablero_Seguimient_Andje\\contratos.xlsx',
            help='Ruta del archivo Excel a importar'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        
        # Mapeo de columnas Excel -> campos del modelo Contract
        COLUMN_MAPPING = {
            "num Total": "num_total",
            "num": "num",
            "Código UNSPSC": "codigo_unspsc",
            "Rubro": "rubro",
            "Descripción del Rubro": "descripcion_rubro",
            "Grupo": "grupo",
            "Subgrupo": "subgrupo",
            "Detalle": "detalle",
            "Pertenece a la OASTI": "pertenece_a_la_oasti",
            "Area a la que pertenece": "area_pertenece",
            "Año PAA": "anio_paa",
            "Línea PAA": "linea_paa",
            "Contratado 2024": "contratado_2024",
            "Adjudicado ?": "adjudicado",
            "Supervisor del contrato": "supervisor_contrato",
            "Apoyo de la supervisión": "apoyo_supervision",
            "Fecha de designacion de supervisión": "fecha_designacion_supervision",
            "Impacto": "impacto",
            "Sistema de Información donde esta publicado el proceso": "sistema_publicacion",
            "Nombre del Contratista": "nombre_contratista",
            "Objeto": "objeto",
            "Numero de proceso o de evento de cotización": "numero_proceso",
            "Numero del contrato": "numero_contrato",
            "Fecha de suscripción de contrato": "fecha_suscripcion_contrato",
            "Fecha aprobación póliza": "fecha_aprobacion_poliza",
            "Fecha de inicio de ejecución (acta de inicio)": "fecha_inicio_ejecucion",
            "Fecha Final": "fecha_final",
            "Fecha de inicio de servicio": "fecha_inicio_servicio",
            "Fecha Final del servicio": "fecha_final_servicio",
            "Fuente vigencia": "fuente_vigencia",
            "Modalida presentación General Cuadro": "modalidad_presentacion_general_cuadro",
            "Tipo de proceso": "tipo_proceso",
            "Valor total incial del contrato": "valor_total_inicial_contrato",
            "Valor adiciones/reducciones": "valor_adiciones_reducciones",
            "Valor total final del contrato": "valor_total_final_contrato",
            "Valor antes del 2024": "valor_antes_2024",
            "Valor asignado al 2024": "valor_asignado_2024",
            "Valor despues del 2024": "valor_despues_2024",
            "Valor total asignado al PAA 2024": "valor_total_asignado_paa_2024",
            "Diferencia con lo contratado respecto al PAA": "diferencia_contratado_respecto_paa",
            "Liberaciones de CDP": "liberaciones_cdp",
            "Cantidad de pagos vigencia 2024 con corte al 31 de julio de 2024": "cantidad_pagos_2024_hasta_31_julio",
            "Valor ejecutado vigencia 2024 Corte 31 de Julio presupuesto 2024": "valor_ejecutado_2024_hasta_31_julio",
            "Pago Mensual Aproximado": "pago_mensual_aproximado",
            "Porcentaje de ejecución presupuestal 2024": "porcentaje_ejecucion_2024",
            "Presupuesto pendiente de ejecutar o  liberar en el 2024": "presupuesto_pendiente_2024",
            "Expediente Gestión Documental": "expediente_gestion_documental",
            "URL SECOP II/TVEC": "url_secop_tv",
            "Observaciones": "observaciones",
        }

        # Set para saber qué campos son fechas, decimales, enteros
        DATE_FIELDS = {
            "fecha_designacion_supervision",
            "fecha_suscripcion_contrato",
            "fecha_aprobacion_poliza",
            "fecha_inicio_ejecucion",
            "fecha_final",
            "fecha_inicio_servicio",
            "fecha_final_servicio"
        }

        DECIMAL_FIELDS = {
            "contratado_2024",
            "valor_total_inicial_contrato",
            "valor_adiciones_reducciones",
            "valor_total_final_contrato",
            "valor_antes_2024",
            "valor_asignado_2024",
            "valor_despues_2024",
            "valor_total_asignado_paa_2024",
            "diferencia_contratado_respecto_paa",
            "liberaciones_cdp",
            "valor_ejecutado_2024_hasta_31_julio",
            "pago_mensual_aproximado",
            "porcentaje_ejecucion_2024",
            "presupuesto_pendiente_2024"
        }

        INT_FIELDS = {
            "num_total",
            "num",
            "anio_paa",
            "cantidad_pagos_2024_hasta_31_julio"
        }

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"No se pudo abrir el archivo {file_path}: {e}")

        if "Contratos 2024" not in wb.sheetnames:
            raise CommandError("La hoja 'Contratos 2024' no existe en el archivo Excel.")

        ws = wb["Contratos 2024"]

        # La fila 7 es la de cabeceras
        header_row = 7
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            val = str(cell.value).strip() if cell.value else ""
            headers[col_idx] = val

        # Verificamos que todas las columnas esperadas estén
        for col_name in COLUMN_MAPPING.keys():
            if col_name not in headers.values():
                self.stderr.write(self.style.WARNING(f"La columna '{col_name}' no se encontró en el Excel."))

        # Iteramos las filas a partir de la 8 en adelante
        created_count = 0
        for row_idx in range(header_row+1, ws.max_row+1):
            row_cells = ws[row_idx]
            # Si la primera celda de la fila está vacía, puede que sea el final
            if all([c.value is None for c in row_cells]):
                continue  # omite filas completamente vacías

            data = {}
            for col_idx, header_name in headers.items():
                if header_name in COLUMN_MAPPING:
                    field_name = COLUMN_MAPPING[header_name]
                    cell_val = ws.cell(row=row_idx, column=col_idx).value

                    # Convertimos según el tipo de campo:
                    if field_name in DATE_FIELDS and cell_val:
                        if isinstance(cell_val, datetime.date):
                            data[field_name] = cell_val
                        else:
                            # Intentamos parsear manualmente si no es date
                            # (si no viene como date)
                            try:
                                data[field_name] = datetime.datetime.strptime(str(cell_val), '%Y-%m-%d').date()
                            except:
                                data[field_name] = None
                    elif field_name in DECIMAL_FIELDS and cell_val is not None:
                        try:
                            data[field_name] = Decimal(str(cell_val).replace(',', '.'))
                        except:
                            data[field_name] = None
                    elif field_name in INT_FIELDS and cell_val is not None:
                        try:
                            data[field_name] = int(cell_val)
                        except:
                            data[field_name] = None
                    else:
                        # Campo texto u otro
                        if cell_val is not None:
                            data[field_name] = str(cell_val).strip()
                        else:
                            data[field_name] = None

            # Creamos el objeto Contract si hay al menos algo en data
            # (en general, sí lo habrá)
            contract = Contract(**data)
            contract.save()
            created_count += 1

        self.stdout.write(self.style.SUCCESS(f"Importación completada. Se crearon {created_count} contratos."))
