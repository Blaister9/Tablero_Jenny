import json
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Configuración del logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def cargar_json(ruta_archivo):
    """Carga datos desde un archivo JSON."""
    try:
        logging.info(f"Cargando JSON desde {ruta_archivo}")
        with open(ruta_archivo, 'r', encoding='utf-8') as file:
            data = json.load(file)
        logging.info("JSON cargado correctamente")
        return data
    except FileNotFoundError:
        logging.error(f"Error: Archivo no encontrado en la ruta {ruta_archivo}")
        return None
    except json.JSONDecodeError:
        logging.error(f"Error: No se pudo decodificar el JSON desde {ruta_archivo}")
        return None
    except Exception as e:
        logging.error(f"Error inesperado al cargar el JSON: {e}")
        return None

def crear_workbook():
    """Crea un nuevo libro de Excel."""
    logging.info("Creando un nuevo libro de Excel")
    workbook = openpyxl.Workbook()
    # Eliminar la hoja por defecto para empezar con un libro vacío
    del workbook['Sheet']
    return workbook

def agregar_hoja(workbook, nombre_hoja):
    """Agrega una nueva hoja al libro de Excel y devuelve el objeto hoja."""
    try:
        logging.info(f"Creando hoja: {nombre_hoja}")
        hoja = workbook.create_sheet(title=nombre_hoja)
        return hoja
    except openpyxl.utils.exceptions.InvalidSheetNameException as e:
        logging.error(f"Error: Nombre de hoja inválido '{nombre_hoja}'. Detalles: {e}")
        return None

def escribir_encabezados(hoja, encabezados):
    """Escribe los encabezados en la hoja de Excel."""
    try:
        logging.info(f"Escribiendo encabezados: {encabezados} en la hoja {hoja.title}")
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            hoja[f"{col_letra}1"] = encabezado
            hoja[f"{col_letra}1"].font = Font(bold=True)  # Estilo negrita
    except Exception as e:
        logging.error(f"Error al escribir encabezados: {e}")

def procesar_nodo(hoja, nodo, ruta_actual="", fila_num=2):
    """
    Procesa recursivamente un nodo del JSON para escribir datos en la hoja de Excel.

    Args:
        hoja: Objeto de la hoja de Excel donde se escribirán los datos.
        nodo: El nodo actual del JSON a procesar (puede ser un diccionario, lista o valor).
        ruta_actual: Ruta actual en el JSON para rastrear la jerarquía.
        fila_num: Número de fila actual en la hoja de Excel.

    Returns:
        El siguiente número de fila disponible en la hoja de Excel.
    """
    logging.debug(f"Procesando nodo en: {ruta_actual}")

    if isinstance(nodo, dict):
        if "preguntas" in nodo:
            # Procesar la lista de preguntas y respuestas
            for pregunta_data in nodo["preguntas"]:
                if isinstance(pregunta_data, dict):
                    pregunta = pregunta_data.get("pregunta", "")
                    respuesta = pregunta_data.get("respuesta", "")
                    
                    # Escribir pregunta y respuesta en la fila actual
                    hoja[f"A{fila_num}"] = ruta_actual  # Usar ruta como "tema"
                    hoja[f"B{fila_num}"] = pregunta
                    hoja[f"C{fila_num}"] = respuesta
                    fila_num += 1  # Incrementar el número de fila

                else:
                    logging.warning(f"Formato inesperado en la lista de preguntas en {ruta_actual}: {type(pregunta_data)}")
        else:
            # Procesar los elementos del diccionario
            for clave, valor in nodo.items():
                nueva_ruta = f"{ruta_actual} > {clave}" if ruta_actual else clave
                fila_num = procesar_nodo(hoja, valor, nueva_ruta, fila_num)  # Pasar fila_num
    elif isinstance(nodo, list):
        # Procesar los elementos de la lista
        for item in nodo:
            fila_num = procesar_nodo(hoja, item, ruta_actual, fila_num)  # Pasar fila_num
    else:
        # Escribir el valor si es un tipo de dato simple
        try:
            hoja[f"A{fila_num}"] = ruta_actual
            hoja[f"B{fila_num}"] = nodo
            fila_num += 1
        except Exception as e:
            logging.warning(f"Error al escribir valor en {ruta_actual}: {e}")

    return fila_num  # Devolver el número de fila actualizado

def json_a_excel(json_data, nombre_archivo="Temas_Juridicos_ANDJE.xlsx"):
    """
    Convierte datos de un JSON a un archivo de Excel, creando una hoja por cada tema principal.
    """
    logging.info("Iniciando procesamiento del JSON")

    workbook = crear_workbook()
    if not workbook:
        logging.error("No se pudo crear el libro de Excel. Saliendo.")
        return

    temas_principales = json_data.get("Temas Principales", {})
    if not isinstance(temas_principales, dict):
        logging.error("Error: 'Temas Principales' no es un diccionario. Saliendo.")
        return

    for tema, contenido_tema in temas_principales.items():
        if not isinstance(tema, str):
            logging.warning(f"Tema con formato inesperado: {type(tema)}")
            continue

        logging.info(f"Procesando tema: {tema}")
        hoja = agregar_hoja(workbook, tema)
        if not hoja:
            logging.warning(f"No se pudo crear la hoja para el tema: {tema}")
            continue

        # Definir encabezados
        encabezados = ["Tema", "Pregunta", "Respuesta"]
        escribir_encabezados(hoja, encabezados)

        # Procesar el contenido del tema
        fila_num = procesar_nodo(hoja, contenido_tema, tema, 2)

        # Ajustar el ancho de las columnas A, B y C
        for col_num in range(1, 4):  # Columnas A, B y C
            col_letra = get_column_letter(col_num)
            hoja.column_dimensions[col_letra].width = 50

        # Verificar si la hoja tiene datos
        if hoja.max_row <= 1:
            logging.warning(f"Hoja {tema} sin datos")
        else:
            logging.info(f"Hoja {tema} creada con {hoja.max_row - 1} registros")  # Contar registros excluyendo encabezado

    try:
        logging.info("Generando archivo Excel")
        workbook.save(nombre_archivo)
        logging.info(f"Archivo Excel generado correctamente: {nombre_archivo}")
    except Exception as e:
        logging.error(f"Error al guardar el archivo Excel: {e}")

# Ejemplo de uso
if __name__ == "__main__":
    ruta_json = r'C:\Users\edwin.paz\Documents\Laura_Chatbot\chatbot_andje_completo.json'  # Reemplaza con la ruta correcta a tu archivo JSON
    data = cargar_json(ruta_json)
    if data:
        json_a_excel(data)
    else:
        logging.error("No se pudo cargar el JSON. Saliendo.")